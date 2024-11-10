from django.views.generic import View, CreateView, UpdateView, ListView, DeleteView
from django.template.loader import render_to_string
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, reverse, HttpResponseRedirect, render
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.contrib import messages

from openpyxl import Workbook, utils as OpenPyXlUtils
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.image import Image
from tempfile import NamedTemporaryFile
from random import shuffle


from visitors.models import SchoolGroup
from areas.models import Area
from .forms import WorksheetForm
from .models import Worksheet, TaskType, Task, Question, Option

import os
import re
import logging
logger = logging.getLogger(__name__)


class WorksheetListView(ListView):
    model = Worksheet
    template_name = 'worksheets_list.html'

    def dispatch(self, request, *args, **kwargs):
        self.area = get_object_or_404(Area, pk=kwargs['area'])

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['area'] = self.area
        context['worksheets'] = Worksheet.objects.filter(area=self.area)

        return context

class BaseWorksheetView(View):
    template_name = 'worksheets_form.html'
    form_class = WorksheetForm
    model = Worksheet

    def dispatch(self, request, *args, **kwargs):
        self.area = get_object_or_404(Area, pk=kwargs['area'])

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['area'] = self.area
        context['school_groups'] = SchoolGroup.objects.all()
        context['task_types'] = TaskType.objects.all()

        return context

    def form_valid(self, form):
        return HttpResponseRedirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, "Formulář obsahuje chyby, prosím opravte je a zkuste znovu.")
        return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        return reverse('worksheets:worksheets_list', kwargs={'area': self.area.pk})

    def create_type_1_task(self, request, task_num):

        task = Task.objects.create(
            worksheet=self.worksheet,
            type=TaskType.objects.get(type=TaskType.Type.TWO_CHOICES),
            text=request.POST.get(f'task-{task_num}-text')
        )

        option_text_0 = request.POST.get(f'option_{task_num}-0-text')
        option_text_1 = request.POST.get(f'option_{task_num}-1-text')

        questions = {k:v for (k, v) in request.POST.items() if k.startswith(f'question_{task_num}') and k.endswith('text')}
        for question_name, question_text in questions.items():

            question_num = re.findall(r'\d+', question_name)[1]

            question = Question.objects.create(
                task=task,
                text=question_text
            )

            is_correct_0 = request.POST.get(f'option_{task_num}-{question_num}-is_correct') == 'is_correct_0'

            Option.objects.bulk_create([
                Option(question=question, text=option_text_0, is_correct=is_correct_0),
                Option(question=question, text=option_text_1, is_correct=not is_correct_0)
            ])

    def create_type_2_or_3_task(self, request, task_num, image=False):

        if image:
            image_src = request.FILES.get(f'{task_num}-image')
            if not image_src:
                image_src = request.POST.get(f'{task_num}-original-image')
                image_src = image_src.replace('/media', '') if image_src else None

            task = Task.objects.create(
                worksheet=self.worksheet,
                type=TaskType.objects.get(type=TaskType.Type.CHOICES_PICTURE),
                text=request.POST.get(f'task-{task_num}-text'),
                image=image_src
            )

            question = Question.objects.create(
                task=task
            )

        else:
            task = Task.objects.create(
                worksheet=self.worksheet,
                type=TaskType.objects.get(type=TaskType.Type.CHOICES),
                text=request.POST.get(f'task-{task_num}-text')
            )

            question = Question.objects.create(
                task=task,
                text=request.POST.get(f'question_{task_num}-text')
            )

        options = {k:v for (k, v) in request.POST.items() if k.startswith(f'option_{task_num}') and k.endswith('text')}

        for option_name, option_text in options.items():
            option_num = re.findall(r'\d+', option_name)[1]
            is_correct = True if request.POST.get(f'option_{task_num}-{option_num}-is_correct') else False

            Option.objects.create(
                question=question,
                text=option_text,
                is_correct=is_correct
            )

    def create_type_4_task(self, request, task_num):

        image_src = request.FILES.get(f'{task_num}-image')
        if not image_src:
            image_src = request.POST.get(f'{task_num}-original-image')
            image_src = image_src.replace('/media', '') if image_src else None

        logger.warning('image_src: %s', image_src)

        task = Task.objects.create(
            worksheet=self.worksheet,
            type=TaskType.objects.get(type=TaskType.Type.MULTIPLE_CHOICES_PICTURE),
            text=request.POST.get(f'task-{task_num}-text'),
            image=image_src
        )

        options = [v for (k, v) in request.POST.items() if k.startswith(f'option_{task_num}') and k.endswith('text')]
        options_correct = [v for (k, v) in request.POST.items() if k.startswith(f'option_{task_num}') and k.endswith('is_correct')]

        for i, correct_option in enumerate(options_correct):
            if not correct_option:
                continue

            question = Question.objects.create(
                task=task,
                text=correct_option
            )

            for j, option in enumerate(options):
                # Correct if number of correct answer is same as number of option
                is_correct = option == options[i]

                Option.objects.create(
                    question=question,
                    text=option,
                    is_correct=is_correct
                )

    def create_type_5_task(self, request, task_num):

        task = Task.objects.create(
            worksheet=self.worksheet,
            type=TaskType.objects.get(type=TaskType.Type.PAIRS),
            text=request.POST.get(f'task-{task_num}-text')
        )

        questions = {k: v for (k, v) in request.POST.items() if k.startswith(f'question_{task_num}') and k.endswith('text')}
        options = {k:v for (k, v) in request.POST.items() if k.startswith(f'option_{task_num}') and k.endswith('text')}

        for question_name, question_text in questions.items():
            question_num = re.findall(r'\d+', question_name)[1]

            question = Question.objects.create(
                task=task,
                text=question_text
            )

            for option_name, option_text in options.items():
                is_correct = True if option_name==('option_'+task_num+'-'+question_num+'-text') else False

                Option.objects.create(
                    question=question,
                    text=option_text,
                    is_correct=is_correct
                )

class WorksheetCreateView(BaseWorksheetView, CreateView):

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        form.instance.area = self.area
        self.object = None

        if not form.is_valid():
            return self.form_invalid(form)

        self.worksheet = form.save()

        school_group_ids = request.POST.getlist('school_group')
        self.worksheet.school_groups.set(school_group_ids)
        self.worksheet.save()

        tasks = {k: v for (k, v) in request.POST.items() if k.startswith('task') and k.endswith('type')}

        for task_type, value in tasks.items():
            task_num = re.findall(r'\d+', task_type)[0]

            if value == '1':
                self.create_type_1_task(request, task_num)
            elif value == '2':
                self.create_type_2_or_3_task(request, task_num)
            elif value == '3':
                self.create_type_2_or_3_task(request, task_num, image=True)
            elif value == '4':
                self.create_type_4_task(request, task_num)
            elif value == '5':
                self.create_type_5_task(request, task_num)

        return super().post(request, *args, **kwargs)


class WorksheetUpdateView(BaseWorksheetView, UpdateView):

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        worksheet = self.get_object()
        context['tasks_data'] = [
            {
                'type': task.type.id,
                'text': task.text or '',
                'image': task.image.url if task.image else '',
                'questions': self.prepare_task_data(task),
            }
            for task in worksheet.task_set.all()
        ]

        context['selected_school_groups'] = list(worksheet.school_groups.all().values_list('pk', flat=True))

        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        form.instance.area = self.area

        self.worksheet = self.get_object()
        form.instance = self.worksheet

        if not form.is_valid():
            return self.form_invalid(form)

        form.save()

        school_group_ids = request.POST.getlist('school_group')
        self.worksheet.school_groups.set(school_group_ids)
        self.worksheet.save()

        # delete all tasks
        self.worksheet.task_set.all().delete()

        logger.warning('post: %s', request.POST)

        tasks = {k: v for (k, v) in request.POST.items() if k.startswith('task') and k.endswith('type')}
        logger.warning('tasks: %s', tasks)

        for task_type, value in tasks.items():
            task_num = re.findall(r'\d+', task_type)[0]

            if value == '1':
                self.create_type_1_task(request, task_num)
            elif value == '2':
                self.create_type_2_or_3_task(request, task_num)
            elif value == '3':
                self.create_type_2_or_3_task(request, task_num, image=True)
            elif value == '4':
                self.create_type_4_task(request, task_num)
            elif value == '5':
                self.create_type_5_task(request, task_num)

        return super().post(request, *args, **kwargs)

    def prepare_task_data(self, task):

        if task.type.id == 1:
            questions = self.prepare_type_1_task_data(task)
        elif task.type.id == 2:
            questions = self.prepare_type_2_or_3_task_data(task)
        elif task.type.id == 3:
            questions = self.prepare_type_2_or_3_task_data(task)
        elif task.type.id == 4:
            questions = self.prepare_type_4_task_data(task)
        elif task.type.id == 5:
            questions = self.prepare_type_5_task_data(task)

        return questions if questions else []


    def prepare_type_1_task_data(self, task):
        questions = task.question_set.all()
        options = questions[0].option_set.all() if questions else []

        return {
            'rows': questions.count(),
            'options': [option.text if option.text else "" for option in options],
            'questions': [
                {
                    'text': question.text if question.text else "",
                    'correct': 0 if question.option_set.first().is_correct else 1,
                }
                for question in questions
            ],
        }

    def prepare_type_2_or_3_task_data(self, task):
        question = task.question_set.first()
        options = question.option_set.all() if question else []

        return {
            'rows': options.count(),
            'question': question.text if question.text else '',
            'options': [
                {'text': option.text if option.text else '',
                 'correct': 'on' if option.is_correct else ''}
                for option in options
            ],
        }

    def prepare_type_4_task_data(self, task):

        questions = task.question_set.all()
        options = questions[0].option_set.all() if questions[0] else []

        correct_answers = []
        for question in questions:
            correct_option = question.option_set.filter(is_correct=True).first()

            if correct_option:
                correct_option = correct_option.text

            correct_answers.append((question.text, correct_option))

        return {
            'rows': options.count(),
            'questions': task.question_set.count(),
            'options': [
                {
                    'text': option.text if option.text else '',
                    'correct': [answer[0] for answer in correct_answers if answer[1] == option.text],
                }
                for option in options
            ],
        }

    def prepare_type_5_task_data(self, task):
        questions = task.question_set.all()

        return {
            'rows': questions.count(),
            'questions': [
                {
                    'text': question.text if question.text else '',
                    'correct': question.option_set.filter(is_correct=True).first().text
                    if question.option_set.filter(is_correct=True).exists()
                    else '',
                }
                for question in questions
            ],
        }


class LoadTaskFormView(View):
    template_mapping = {
        '1': 'task_type_1.html',
        '2': 'task_type_2.html',
        '3': 'task_type_3.html',
        '4': 'task_type_4.html',
        '5': 'task_type_5.html',
    }

    def get(self, request, *args, **kwargs):
        task_type = kwargs.get('task_type')
        template_name = self.template_mapping.get(task_type)

        if template_name:
            html_content = render_to_string(template_name, request=request)
            return JsonResponse({'html': html_content})
        else:
            return JsonResponse({'error': 'Invalid task type'}, status=400)


class CheckFormDataAjaxView(View):

    errors = {}

    def post(self, request, *args, **kwargs):
        self.errors = {}
        logger.warning('CheckFormDataAjaxView: %s', request.POST)
        title = request.POST.get('title')

        if not title:
            self.errors['title'] = "Název je povinný."

        tasks = {k: v for (k, v) in request.POST.items() if k.startswith('task') and k.endswith('type')}

        images = {k: v for (k, v) in request.POST.items() if k.endswith('image')}
        for image_name, image in images.items():
            task_num = re.findall(r'\d+', image_name)[0]
            image = image or request.POST.get(f'{task_num}-original-image')

            if not image:
                self.errors[image_name] = 'Obrázek je povinný.'


        questions = {k: v for (k, v) in request.POST.items() if k.startswith('question') and k.endswith('text')}
        for question_name, question in questions.items():
            if not question:
                self.errors[question_name] = 'Toto pole je povinné.'
            elif len(question) >= 100:
                self.errors[question_name] = f'Maximální počet znaků je 150. Máte {len(question)}.'

        options = {k:v for (k, v) in request.POST.items() if k.startswith('option') and k.endswith('text')}
        for option_name, option in options.items():
            if not option:
                self.errors[option_name] = 'Toto pole je povinné.'
            elif len(option) >= 50:
                self.errors[option_name] = f'Maximální počet znaků je 50. Máte {len(option)}.'


        for task_type, value in tasks.items():
            task_num = re.findall(r'\d+', task_type)[0]

            task_text = request.POST.get(f'task-{task_num}-text')
            if not task_text:
                self.errors[f'task-{task_num}-text'] = 'Zadání je povinné.'
            elif len(task_text) >= 150:
                self.errors[f'task-{task_num}-text'] = f'Maximální počet znaků je 150. Máte {len(task_text)}.'

            elif value == '2':
                self.check_type_2_or_3_task_data(request, task_num)
            elif value == '3':
                self.check_type_2_or_3_task_data(request, task_num)
            elif value == '4':
                self.check_type_4_task_data(request, task_num)

        if self.errors:
            return JsonResponse({'status': False, 'errors': self.errors})

        return JsonResponse({'status': True})


    def check_type_2_or_3_task_data(self, request, task_num):

        options_checkboxes = {k: v for (k, v) in request.POST.items() if k.startswith(f'option_{task_num}') and k.endswith('-is_correct')}

        if not options_checkboxes:
            self.errors[f'checkbox-{task_num}'] = 'Vyberte správnou odpověď.'


    def check_type_4_task_data(self, request, task_num):
        counter_value = request.POST.get(f'question_{task_num}-counter')

        if not counter_value.isdigit():
            self.errors[f'question_{task_num}-counter'] = 'Toto pole musí obsahovat číslo.'

        else:
            counter_value = int(counter_value)
            if not (1 <= counter_value <= 15):
                self.errors[f'question_{task_num}-counter'] = 'Číslo musí být v rozmezí od 1 do 15.'

        options_correct_raw = [v for (k, v) in request.POST.items() if k.startswith(f'option_{task_num}') and k.endswith('is_correct')]
        options_correct = []

        for value in options_correct_raw:
            if value.isdigit():
                options_correct.append(int(value))

        logger.warning('check_type_4_task_data: %s', options_correct)

        if len(options_correct) != counter_value:
            logger.warning('check_type_4_task_data: %s', len(options_correct))
            self.errors[f'select-{task_num}'] = 'Musíte použít všechna čísla.'

        if len(options_correct) != len(set(options_correct)):
            self.errors[f'select-{task_num}'] = 'Každé číslo může být použito pouze jednou.'


class WorksheetDeleteView(DeleteView):
    model = Worksheet

    def dispatch(self, request, *args, **kwargs):
        self.area = get_object_or_404(Area, pk=kwargs['area'])

        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('worksheets:worksheets_list', kwargs={'area': self.area.pk})


class BaseWorksheetExportView(View):

    worksheet = None
    sheet = None
    row_cnt = 0
    task_cnt = 0

    def dispatch(self, request, *args, **kwargs):
        self.worksheet = get_object_or_404(Worksheet, pk=kwargs['pk'])

        return super().dispatch(request, *args, **kwargs)

    def change_column_width(self, column_widths):
        for i, width in enumerate(column_widths):
            column_letter = OpenPyXlUtils.get_column_letter(i + 1)
            self.sheet.column_dimensions[column_letter].width = width

    def make_header(self):
        self.sheet.merge_cells('A1:O3')
        self.sheet['A1'] = self.worksheet.title
        self.sheet['A1'].font = Font(name='Calibri', bold=True, size=20)
        self.sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        self.row_cnt += 5

        self.sheet.merge_cells(f'A{self.row_cnt}:C{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = 'Jméno: '
        self.sheet[f'A{self.row_cnt}'].alignment = Alignment(horizontal='right')

        self.sheet.merge_cells(f'D{self.row_cnt}:M{self.row_cnt}')
        self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

        self.row_cnt += 2

        self.sheet.merge_cells(f'A{self.row_cnt}:C{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = 'Datum: '
        self.sheet[f'A{self.row_cnt}'].alignment = Alignment(horizontal='right')

        self.sheet.merge_cells(f'D{self.row_cnt}:G{self.row_cnt}')
        self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

        self.sheet.merge_cells(f'H{self.row_cnt}:I{self.row_cnt}')
        self.sheet[f'H{self.row_cnt}'] = 'Třída:  '
        self.sheet[f'H{self.row_cnt}'].alignment = Alignment(horizontal='right')

        self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
        self.sheet[f'J{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

        self.row_cnt += 3


    def make_task_text(self, task_num, text):
        self.sheet.merge_cells(f'A{self.row_cnt}:O{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = f'{task_num}. {text}'
        self.sheet[f'A{self.row_cnt}'].font = Font(name='Calibri', bold=True)
        self.row_cnt += 2

    def add_checkbox_image(self, img_path, col_start):

        img = Image(img_path)

        cell_width = 99200
        cell_height = 75600

        offset_x = int((cell_width - cell_height) / 2)

        _from = AnchorMarker(col=col_start, row=self.row_cnt - 1, colOff=offset_x, rowOff=0)
        to = AnchorMarker(col=col_start + 1, row=self.row_cnt, colOff=-offset_x, rowOff=0)
        img.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to)

        self.sheet.add_image(img)

    def add_image(self, img_path):

        img = Image(img_path)

        cell_height = 75600
        max_width =  1688400

        img_width = img.width
        img_height = img.height

        new_height = (img_height * max_width) / img_width
        offset_x = 0

        # If image is higher than 15 cells
        if new_height > 15 * cell_height:
            new_height = 15 * cell_height
            new_width = (img_width * new_height) / img_height

            offset_x = int((max_width - new_width) * 1.4)

        rows = int(new_height / cell_height)
        offset_y = int((new_height % cell_height) / 2)

        _from = AnchorMarker(col=2, row=self.row_cnt - 1, colOff=offset_x, rowOff=offset_y)
        to = AnchorMarker(col=13, row=self.row_cnt + rows, colOff=-offset_x, rowOff=-offset_y)
        img.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to)

        self.shift_task_to_next_page(rows + 3)

        self.sheet.add_image(img)
        self.row_cnt += rows + 2

    def shift_task_to_next_page(self, task_rows):
        rows_per_page = 50

        # if at least one half of task is in another page
        if (self.row_cnt % rows_per_page) + (task_rows / 2) > rows_per_page:

            rows = rows_per_page - (self.row_cnt % rows_per_page)
            self.row_cnt += rows


    def make_two_choices_task(self, task):

        questions = task.question_set.all()
        options = questions[0].option_set.all() if len(questions) else []

        # Shift task to another page if task is too long
        task_rows = 4 + (2 * len(questions))
        self.shift_task_to_next_page(task_rows)

        self.make_task_text(self.task_cnt + 1, task.text)

        self.sheet.merge_cells(f'F{self.row_cnt}:J{self.row_cnt}')
        self.sheet[f'F{self.row_cnt}'] = options[0].text
        self.sheet[f'F{self.row_cnt}'].alignment = Alignment(horizontal='center')

        self.sheet.merge_cells(f'K{self.row_cnt}:M{self.row_cnt}')
        self.sheet[f'K{self.row_cnt}'] = options[1].text
        self.sheet[f'K{self.row_cnt}'].alignment = Alignment(horizontal='center')
        self.row_cnt += 1

        self.sheet.row_dimensions[self.row_cnt].height = 10
        self.row_cnt += 1

        return questions


    def make_choices_task(self, task):

        question = task.question_set.first()
        options = question.option_set.all()

        # Shift task to another page if task is too long
        task_rows = 4 + (2 * len(question.option_set.all()))
        self.shift_task_to_next_page(task_rows)

        self.make_task_text(self.task_cnt + 1, task.text)

        self.sheet.merge_cells(f'B{self.row_cnt}:O{self.row_cnt}')
        self.sheet[f'B{self.row_cnt}'] = question.text
        self.row_cnt += 1

        self.sheet.row_dimensions[self.row_cnt].height = 10
        self.row_cnt += 1

        return options


    def make_choices_picture_task(self, task):

        question = task.question_set.first()
        options = question.option_set.all()

        self.make_task_text(self.task_cnt + 1, task.text)

        if task.image:
            self.add_image(task.image.path)

        # Shift task to another page if task is too long
        task_rows = 1 + (2 * len(options))
        self.shift_task_to_next_page(task_rows)

        return options


    def make_multiple_choices_picture_task(self, task):

        question = task.question_set.first()

        self.make_task_text(self.task_cnt + 1, task.text)

        if task.image:
            self.add_image(task.image.path)

        all_options = [option.text for option in question.option_set.all()]

        self.sheet.merge_cells(f'A{self.row_cnt}:O{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = f'Výběr pojmů: {", ".join(all_options)}'
        self.row_cnt += 2

        return all_options

    def make_pairs_task(self, task):

        questions = task.question_set.all()
        options = list(questions[0].option_set.all() if len(questions) else [])

        # Shift task to another page if task is too long
        task_rows = 2 + (2 * len(questions))
        self.shift_task_to_next_page(task_rows)

        self.make_task_text(self.task_cnt + 1, task.text)

        return questions, options


    def post(self, request, **kwargs):

        workbook = Workbook()
        self.sheet = workbook.active
        self.change_column_width([3, 3, 3, 9, 6, 6, 3, 3, 6, 3, 9, 3, 9, 3, 3])
        self.make_header()

        tasks = self.worksheet.task_set.all()
        for task_num, task in enumerate(tasks):
            self.task_cnt = task_num + 1

            if task.type.type == TaskType.Type.TWO_CHOICES:
                self.make_two_choices_task(task)

            elif task.type.type == TaskType.Type.CHOICES:
                self.make_choices_task(task)

            elif task.type.type == TaskType.Type.CHOICES_PICTURE:
                self.make_choices_picture_task(task)

            elif task.type.type == TaskType.Type.MULTIPLE_CHOICES_PICTURE:
                self.make_multiple_choices_picture_task(task)

            elif task.type.type == TaskType.Type.PAIRS:
                self.make_pairs_task(task)

            self.row_cnt += 2

        self.row_cnt += 1

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Access-Control-Expose-Headers'] = f'Content-Disposition'
        response['Content-Disposition'] = f'attachment; filename=pracovni_list.xlsx'
        return response


@method_decorator(csrf_exempt, name='dispatch')
class WorksheetExportView(BaseWorksheetExportView):

    def make_two_choices_task(self, task):

        questions = super().make_two_choices_task(task)

        img_path = os.path.join(settings.STATIC_ROOT, 'box.jpeg')

        for question in questions:
            self.sheet.merge_cells(f'B{self.row_cnt}:E{self.row_cnt}')
            self.sheet[f'B{self.row_cnt}'] = question.text

            self.add_checkbox_image(img_path, col_start=7)
            self.add_checkbox_image(img_path, col_start=11)

            self.row_cnt += 1
            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1

    def make_choices_task(self, task):

        options = super().make_choices_task(task)

        char = ord('a')
        for option in options:
            self.sheet[f'C{self.row_cnt}'] = f'{chr(char)})'
            self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
            char += 1

            self.sheet.merge_cells(f'D{self.row_cnt}:O{self.row_cnt}')
            self.sheet[f'D{self.row_cnt}'] = option.text
            self.row_cnt += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_choices_picture_task(self, task):

        options = super().make_choices_picture_task(task)

        char = ord('a')
        for option in options:
            self.sheet.merge_cells(f'E{self.row_cnt}:M{self.row_cnt}')
            self.sheet[f'E{self.row_cnt}'] = f'{chr(char)}) {option.text}'
            self.row_cnt += 1
            char += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_multiple_choices_picture_task(self, task):

        all_options = super().make_multiple_choices_picture_task(task)

        for i in range(len(all_options)):

            if i % 2 == 0:
                self.sheet[f'C{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'D{self.row_cnt}:G{self.row_cnt}')
                self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

            else:
                self.sheet[f'I{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'I{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
                self.sheet[f'J{self.row_cnt}'].border = Border(bottom=Side(style='thin'))
                self.row_cnt += 1


                self.sheet.row_dimensions[self.row_cnt].height = 10
                self.row_cnt += 1

        self.row_cnt += 1


    def make_pairs_task(self, task):

        questions, options = super().make_pairs_task(task)

        shuffle(options)

        for i, question in enumerate(questions):
            self.sheet.merge_cells(f'C{self.row_cnt}:F{self.row_cnt}')
            self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
            self.sheet[f'C{self.row_cnt}'] = question.text

            self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
            self.sheet[f'J{self.row_cnt}'].alignment = Alignment(horizontal='center')
            self.sheet[f'J{self.row_cnt}'] = options[i].text

            self.row_cnt += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


@method_decorator(csrf_exempt, name='dispatch')
class WorksheetExportWithAnswersView(BaseWorksheetExportView):

    def make_two_choices_task(self, task):

        questions = super().make_two_choices_task(task)

        box_path = os.path.join(settings.STATIC_ROOT, 'box.jpeg')
        checked_box_path = os.path.join(settings.STATIC_ROOT, 'checked_box.jpeg')

        for question in questions:
            self.sheet.merge_cells(f'B{self.row_cnt}:E{self.row_cnt}')
            self.sheet[f'B{self.row_cnt}'] = question.text

            options = question.option_set.all()
            correct_option = options.filter(is_correct=True).first()

            if correct_option.text == options[0].text:
                self.add_checkbox_image(box_path, col_start=7)
                self.add_checkbox_image(checked_box_path, col_start=11)
            else:
                self.add_checkbox_image(box_path, col_start=11)
                self.add_checkbox_image(checked_box_path, col_start=7)

            self.row_cnt += 1
            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_choices_task(self, task):

        options = super().make_choices_task(task)

        char = ord('a')
        for option in options:

            self.sheet[f'C{self.row_cnt}'] = f'{chr(char)})'
            self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
            char += 1

            self.sheet.merge_cells(f'D{self.row_cnt}:O{self.row_cnt}')
            self.sheet[f'D{self.row_cnt}'] = option.text

            if option.is_correct:
                self.sheet[f'C{self.row_cnt}'].font = Font(name='Calibri', bold=True)
                self.sheet[f'D{self.row_cnt}'].font = Font(name='Calibri', bold=True)

            self.row_cnt += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_choices_picture_task(self, task):

        options = super().make_choices_picture_task(task)

        char = ord('a')
        for option in options:
            self.sheet.merge_cells(f'E{self.row_cnt}:M{self.row_cnt}')
            self.sheet[f'E{self.row_cnt}'] = f'{chr(char)}) {option.text}'

            if option.is_correct:
                self.sheet[f'E{self.row_cnt}'].font = Font(name='Calibri', bold=True)

            self.row_cnt += 1
            char += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_multiple_choices_picture_task(self, task):

        all_options = super().make_multiple_choices_picture_task(task)
        questions = task.question_set.all()

        for i in range(len(all_options)):

            question = questions.filter(text=i + 1).first()
            correct_answer = question.option_set.get(is_correct=True).text if question.option_set.get(is_correct=True) else ''

            if i % 2 == 0:
                self.sheet[f'C{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'D{self.row_cnt}:G{self.row_cnt}')
                self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))
                self.sheet[f'D{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'D{self.row_cnt}'] = correct_answer

            else:
                self.sheet[f'I{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'I{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
                self.sheet[f'J{self.row_cnt}'].border = Border(bottom=Side(style='thin'))
                self.sheet[f'J{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'J{self.row_cnt}'] = correct_answer

                self.row_cnt += 1


                self.sheet.row_dimensions[self.row_cnt].height = 10
                self.row_cnt += 1

            self.row_cnt += 1


    def make_pairs_task(self, task):

            questions, options = super().make_pairs_task(task)

            for i, question in enumerate(questions):
                self.sheet.merge_cells(f'C{self.row_cnt}:F{self.row_cnt}')
                self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'C{self.row_cnt}'] = question.text

                self.sheet.merge_cells(f'G{self.row_cnt}:I{self.row_cnt}')
                self.sheet[f'G{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'G{self.row_cnt}'] = '-'

                self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
                self.sheet[f'J{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'J{self.row_cnt}'] = options[i].text

                self.row_cnt += 1

                self.sheet.row_dimensions[self.row_cnt].height = 10
                self.row_cnt += 1