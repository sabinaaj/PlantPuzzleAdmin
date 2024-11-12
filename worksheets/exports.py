from django.views.generic import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponse

from openpyxl import Workbook, utils as OpenPyXlUtils
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.drawing.image import Image

from random import shuffle
import os

from django.shortcuts import render
from django.http import HttpResponse
from weasyprint import HTML
from io import BytesIO


from .models import Worksheet, TaskType

class BaseWorksheetExportView(View):

    worksheet = None
    sheet = None
    row_cnt = 0
    task_cnt = 0

    def dispatch(self, request, *args, **kwargs):
        self.worksheet = get_object_or_404(Worksheet, pk=kwargs['pk'])

        return super().dispatch(request, *args, **kwargs)

    def change_column_width(self, column_widths):
        for i, width in enumerate(column_widths):
            column_letter = OpenPyXlUtils.get_column_letter(i + 1)
            self.sheet.column_dimensions[column_letter].width = width

    def make_header(self):
        self.sheet.merge_cells('A1:O3')
        self.sheet['A1'] = self.worksheet.title
        self.sheet['A1'].font = Font(name='Calibri', bold=True, size=20)
        self.sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        self.row_cnt += 5

        self.sheet.merge_cells(f'A{self.row_cnt}:C{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = 'Jméno: '
        self.sheet[f'A{self.row_cnt}'].alignment = Alignment(horizontal='right')

        self.sheet.merge_cells(f'D{self.row_cnt}:M{self.row_cnt}')
        self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

        self.row_cnt += 2

        self.sheet.merge_cells(f'A{self.row_cnt}:C{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = 'Datum: '
        self.sheet[f'A{self.row_cnt}'].alignment = Alignment(horizontal='right')

        self.sheet.merge_cells(f'D{self.row_cnt}:G{self.row_cnt}')
        self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

        self.sheet.merge_cells(f'H{self.row_cnt}:I{self.row_cnt}')
        self.sheet[f'H{self.row_cnt}'] = 'Třída:  '
        self.sheet[f'H{self.row_cnt}'].alignment = Alignment(horizontal='right')

        self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
        self.sheet[f'J{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

        self.row_cnt += 3


    def make_task_text(self, task_num, text):
        self.sheet.merge_cells(f'A{self.row_cnt}:O{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = f'{task_num}. {text}'
        self.sheet[f'A{self.row_cnt}'].font = Font(name='Calibri', bold=True)
        self.row_cnt += 2

    def add_checkbox_image(self, img_path, col_start):

        img = Image(img_path)

        cell_width = 99200
        cell_height = 75600

        offset_x = int((cell_width - cell_height) / 2)

        _from = AnchorMarker(col=col_start, row=self.row_cnt - 1, colOff=offset_x, rowOff=0)
        to = AnchorMarker(col=col_start + 1, row=self.row_cnt, colOff=-offset_x, rowOff=0)
        img.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to)

        self.sheet.add_image(img)

    def add_image(self, img_path):

        img = Image(img_path)

        cell_height = 75600
        max_width =  1688400

        img_width = img.width
        img_height = img.height

        new_height = (img_height * max_width) / img_width
        offset_x = 0

        # If image is higher than 15 cells
        if new_height > 15 * cell_height:
            new_height = 15 * cell_height
            new_width = (img_width * new_height) / img_height

            offset_x = int((max_width - new_width) * 1.4)

        rows = int(new_height / cell_height)
        offset_y = int((new_height % cell_height) / 2)

        _from = AnchorMarker(col=2, row=self.row_cnt - 1, colOff=offset_x, rowOff=offset_y)
        to = AnchorMarker(col=13, row=self.row_cnt + rows, colOff=-offset_x, rowOff=-offset_y)
        img.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to)

        self.shift_task_to_next_page(rows + 3)

        self.sheet.add_image(img)
        self.row_cnt += rows + 2

    def shift_task_to_next_page(self, task_rows):
        rows_per_page = 50

        # if at least one half of task is in another page
        if (self.row_cnt % rows_per_page) + (task_rows / 2) > rows_per_page:

            rows = rows_per_page - (self.row_cnt % rows_per_page)
            self.row_cnt += rows


    def make_two_choices_task(self, task):

        questions = task.question_set.all()
        options = questions[0].option_set.all() if len(questions) else []

        # Shift task to another page if task is too long
        task_rows = 4 + (2 * len(questions))
        self.shift_task_to_next_page(task_rows)

        self.make_task_text(self.task_cnt + 1, task.text)

        self.sheet.merge_cells(f'F{self.row_cnt}:J{self.row_cnt}')
        self.sheet[f'F{self.row_cnt}'] = options[0].text
        self.sheet[f'F{self.row_cnt}'].alignment = Alignment(horizontal='center')

        self.sheet.merge_cells(f'K{self.row_cnt}:M{self.row_cnt}')
        self.sheet[f'K{self.row_cnt}'] = options[1].text
        self.sheet[f'K{self.row_cnt}'].alignment = Alignment(horizontal='center')
        self.row_cnt += 1

        self.sheet.row_dimensions[self.row_cnt].height = 10
        self.row_cnt += 1

        return questions


    def make_choices_task(self, task):

        question = task.question_set.first()
        options = question.option_set.all()

        # Shift task to another page if task is too long
        task_rows = 4 + (2 * len(question.option_set.all()))
        self.shift_task_to_next_page(task_rows)

        self.make_task_text(self.task_cnt + 1, task.text)

        self.sheet.merge_cells(f'B{self.row_cnt}:O{self.row_cnt}')
        self.sheet[f'B{self.row_cnt}'] = question.text
        self.row_cnt += 1

        self.sheet.row_dimensions[self.row_cnt].height = 10
        self.row_cnt += 1

        return options


    def make_choices_picture_task(self, task):

        question = task.question_set.first()
        options = question.option_set.all()

        self.make_task_text(self.task_cnt + 1, task.text)

        if task.image:
            self.add_image(task.image.path)

        # Shift task to another page if task is too long
        task_rows = 1 + (2 * len(options))
        self.shift_task_to_next_page(task_rows)

        return options


    def make_multiple_choices_picture_task(self, task):

        question = task.question_set.first()

        self.make_task_text(self.task_cnt + 1, task.text)

        if task.image:
            self.add_image(task.image.path)

        all_options = [option.text for option in question.option_set.all()]

        self.sheet.merge_cells(f'A{self.row_cnt}:O{self.row_cnt}')
        self.sheet[f'A{self.row_cnt}'] = f'Výběr pojmů: {", ".join(all_options)}'
        self.row_cnt += 2

        return all_options

    def make_pairs_task(self, task):

        questions = task.question_set.all()
        options = list(questions[0].option_set.all() if len(questions) else [])

        # Shift task to another page if task is too long
        task_rows = 2 + (2 * len(questions))
        self.shift_task_to_next_page(task_rows)

        self.make_task_text(self.task_cnt + 1, task.text)

        return questions, options

    def post(self, request, **kwargs):
        # Předpokládám, že máte nějakou HTML šablonu, kterou chcete použít pro generování PDF
        tasks = self.worksheet.task_set.all()
        context = {
            'tasks': tasks,
        }

        # Vygenerování HTML z šablony (kde tasks jsou data pro generování tabulky)
        html_content = render(request, 'new_template.html', context)

        # Vytvoření PDF z HTML
        pdf_file = BytesIO()
        HTML(string=html_content.content.decode()).write_pdf(pdf_file)
        pdf_file.seek(0)

        # Vytvoření odpovědi s PDF souborem
        response = HttpResponse(content=pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="pracovni_list.pdf"'

        return response


@method_decorator(csrf_exempt, name='dispatch')
class WorksheetExportView(BaseWorksheetExportView):

    def make_two_choices_task(self, task):

        questions = super().make_two_choices_task(task)

        img_path = os.path.join(settings.STATIC_ROOT, 'box.jpeg')

        for question in questions:
            self.sheet.merge_cells(f'B{self.row_cnt}:E{self.row_cnt}')
            self.sheet[f'B{self.row_cnt}'] = question.text

            self.add_checkbox_image(img_path, col_start=7)
            self.add_checkbox_image(img_path, col_start=11)

            self.row_cnt += 1
            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1

    def make_choices_task(self, task):

        options = super().make_choices_task(task)

        char = ord('a')
        for option in options:
            self.sheet[f'C{self.row_cnt}'] = f'{chr(char)})'
            self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
            char += 1

            self.sheet.merge_cells(f'D{self.row_cnt}:O{self.row_cnt}')
            self.sheet[f'D{self.row_cnt}'] = option.text
            self.row_cnt += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_choices_picture_task(self, task):

        options = super().make_choices_picture_task(task)

        char = ord('a')
        for option in options:
            self.sheet.merge_cells(f'E{self.row_cnt}:M{self.row_cnt}')
            self.sheet[f'E{self.row_cnt}'] = f'{chr(char)}) {option.text}'
            self.row_cnt += 1
            char += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_multiple_choices_picture_task(self, task):

        all_options = super().make_multiple_choices_picture_task(task)

        for i in range(len(all_options)):

            if i % 2 == 0:
                self.sheet[f'C{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'D{self.row_cnt}:G{self.row_cnt}')
                self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))

            else:
                self.sheet[f'I{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'I{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
                self.sheet[f'J{self.row_cnt}'].border = Border(bottom=Side(style='thin'))
                self.row_cnt += 1


                self.sheet.row_dimensions[self.row_cnt].height = 10
                self.row_cnt += 1

        self.row_cnt += 1


    def make_pairs_task(self, task):

        questions, options = super().make_pairs_task(task)

        shuffle(options)

        for i, question in enumerate(questions):
            self.sheet.merge_cells(f'C{self.row_cnt}:F{self.row_cnt}')
            self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
            self.sheet[f'C{self.row_cnt}'] = question.text

            self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
            self.sheet[f'J{self.row_cnt}'].alignment = Alignment(horizontal='center')
            self.sheet[f'J{self.row_cnt}'] = options[i].text

            self.row_cnt += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


@method_decorator(csrf_exempt, name='dispatch')
class WorksheetExportWithAnswersView(BaseWorksheetExportView):

    def make_two_choices_task(self, task):

        questions = super().make_two_choices_task(task)

        box_path = os.path.join(settings.STATIC_ROOT, 'box.jpeg')
        checked_box_path = os.path.join(settings.STATIC_ROOT, 'checked_box.jpeg')

        for question in questions:
            self.sheet.merge_cells(f'B{self.row_cnt}:E{self.row_cnt}')
            self.sheet[f'B{self.row_cnt}'] = question.text

            options = question.option_set.all()
            correct_option = options.filter(is_correct=True).first()

            if correct_option.text == options[0].text:
                self.add_checkbox_image(box_path, col_start=7)
                self.add_checkbox_image(checked_box_path, col_start=11)
            else:
                self.add_checkbox_image(box_path, col_start=11)
                self.add_checkbox_image(checked_box_path, col_start=7)

            self.row_cnt += 1
            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_choices_task(self, task):

        options = super().make_choices_task(task)

        char = ord('a')
        for option in options:

            self.sheet[f'C{self.row_cnt}'] = f'{chr(char)})'
            self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
            char += 1

            self.sheet.merge_cells(f'D{self.row_cnt}:O{self.row_cnt}')
            self.sheet[f'D{self.row_cnt}'] = option.text

            if option.is_correct:
                self.sheet[f'C{self.row_cnt}'].font = Font(name='Calibri', bold=True)
                self.sheet[f'D{self.row_cnt}'].font = Font(name='Calibri', bold=True)

            self.row_cnt += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_choices_picture_task(self, task):

        options = super().make_choices_picture_task(task)

        char = ord('a')
        for option in options:
            self.sheet.merge_cells(f'E{self.row_cnt}:M{self.row_cnt}')
            self.sheet[f'E{self.row_cnt}'] = f'{chr(char)}) {option.text}'

            if option.is_correct:
                self.sheet[f'E{self.row_cnt}'].font = Font(name='Calibri', bold=True)

            self.row_cnt += 1
            char += 1

            self.sheet.row_dimensions[self.row_cnt].height = 10
            self.row_cnt += 1


    def make_multiple_choices_picture_task(self, task):

        all_options = super().make_multiple_choices_picture_task(task)
        questions = task.question_set.all()

        for i in range(len(all_options)):

            question = questions.filter(text=i + 1).first()
            correct_answer = question.option_set.get(is_correct=True).text if question.option_set.get(is_correct=True) else ''

            if i % 2 == 0:
                self.sheet[f'C{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'D{self.row_cnt}:G{self.row_cnt}')
                self.sheet[f'D{self.row_cnt}'].border = Border(bottom=Side(style='thin'))
                self.sheet[f'D{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'D{self.row_cnt}'] = correct_answer

            else:
                self.sheet[f'I{self.row_cnt}'] = f'{i + 1}.  '
                self.sheet[f'I{self.row_cnt}'].alignment = Alignment(horizontal='right')

                self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
                self.sheet[f'J{self.row_cnt}'].border = Border(bottom=Side(style='thin'))
                self.sheet[f'J{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'J{self.row_cnt}'] = correct_answer

                self.row_cnt += 1


                self.sheet.row_dimensions[self.row_cnt].height = 10
                self.row_cnt += 1

            self.row_cnt += 1


    def make_pairs_task(self, task):

            questions, options = super().make_pairs_task(task)

            for i, question in enumerate(questions):
                self.sheet.merge_cells(f'C{self.row_cnt}:F{self.row_cnt}')
                self.sheet[f'C{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'C{self.row_cnt}'] = question.text

                self.sheet.merge_cells(f'G{self.row_cnt}:I{self.row_cnt}')
                self.sheet[f'G{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'G{self.row_cnt}'] = '-'

                self.sheet.merge_cells(f'J{self.row_cnt}:M{self.row_cnt}')
                self.sheet[f'J{self.row_cnt}'].alignment = Alignment(horizontal='center')
                self.sheet[f'J{self.row_cnt}'] = options[i].text

                self.row_cnt += 1

                self.sheet.row_dimensions[self.row_cnt].height = 10
                self.row_cnt += 1

